#!/usr/bin/python
# -*- coding: UTF-8 -*-
from appium import webdriver
from TKIOSAutoDownLoadApp.baseConfg.tk_baseIosPhone import *
from selenium.webdriver.common.by import By
from appium.webdriver.common.appiumby import AppiumBy
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as exc
import unittest
from openpyxl import load_workbook, Workbook
from TKIOSAutoDownLoadApp.proxy_ip import ProxyIp, IpInfo
import requests
import time
import random
from TKIOSAutoDownLoadApp.config import getConfig, TK_Config


class MyDesiredCapabilities:

    def __init__(self):
        self.config: TK_Config = getConfig()

    def get_desired_capabilities(self, udid, port, wdaPort):
        desired_caps = {
            # 平台名称
            'platformName': "iOS",
            # 平台版本
            'platformVersion': get_ios_version(udid),
            # 设备名称 li的iPhone
            'deviceName': get_ios_product_name(udid),
            # 如果填bundlid会无法调起应用
            'app': 'com.tk.getIdfa.TKIDFA',
            # 设备UDID
            'udid': udid,
            # 是否不重新安装启动
            'noReset': True,
            # 超时时间
            'newCommandTimeout': 6000,
            # 自动化测试平台
            'automationName': 'XCUITest',
            # "xcodeSigningId": "iPhone Developer",
            # 'xcodeOrgId': "XH3Y936B53",
            'autoLaunch': True,
            'clearSystemFiles': True,
            # 'showIOSLog': True,
            'wdaLocalPort': wdaPort

        }
        print(desired_caps)
        remote_url = 'http://localhost:' + str(port) + '/wd/hub'
        print(remote_url)
        driver = webdriver.Remote(remote_url, desired_caps)
        self.startTest(driver)
        # self.other_test(driver)
        return driver

    def other_test(self, driver: webdriver):
        pass

    def startTest(self, driver: webdriver):

        app_is_exited = driver.is_app_installed(self.config.get_app_bundle_id())
        if app_is_exited:
            # app 存在就删除
            print("[%s]目标app存在" % driver.capabilities["deviceName"])
            driver.remove_app(self.config.get_app_bundle_id())
            print("[%s]已删除目标APP重新下载" % driver.capabilities["deviceName"])
            self.downAppTest_01(driver)
        else:
            print("[%s]目标APP不存在" % driver.capabilities["deviceName"])
            self.downAppTest_01(driver)

    def downAppTest_01(self, driver: webdriver):

        idfaLab = driver.find_element(by=By.NAME, value='idfalabIdentifier')
        idfaStr = idfaLab.text

        proxy_ip = ProxyIp(maxsize=1, threshold=1)
        ip_info: IpInfo = proxy_ip.get_proxy_ip_info()
        ip_address = ip_info.get_ip()
        ip_port = ip_info.get_port()

        app_report_flag: bool = self.config.get_app_report_flag()
        r_status = False
        if app_report_flag == True:
            # 上报
            url = self.config.get_app_report_link() % (idfaStr, ip_address)
            print("%s上报链接:%s" % (driver.capabilities["deviceName"], url))
            r = requests.get(url)
            if r.status_code == 200:
                r_status = True
                print("[%s]上报成功" % driver.capabilities["deviceName"])

        cleanBtn = driver.find_element(by=By.NAME, value="Clear text")
        cleanBtn.click()
        tk_httpFiled = driver.find_element(by=By.NAME, value='tk_httpFiled')
        tk_httpFiled.send_keys(self.config.get_downApplink())
        time.sleep(2)
        driver.find_element(by=By.NAME, value="tk_autoDownloadApp").click()
        time.sleep(6)
        driver.find_element(by=By.NAME, value="重新下载").click()
        print("[%s]正在下载中,请稍后" % driver.capabilities["deviceName"])
        # 判断app是否存在
        while True:
            time.sleep(2)
            if driver.is_app_installed(self.config.get_app_bundle_id()):
                print("[%s]目标APP下载完成" % driver.capabilities["deviceName"])
                break
        driver.background_app(-1)

        # 配置代理
        self.proxyIp_Change(driver, ip_address, ip_port)
        # 打开目标APP
        driver.activate_app(self.config.get_app_bundle_id())
        print("[%s]已打开目标APP" % driver.capabilities["deviceName"])
        # 处理权限
        self.always_allow(driver)
        open_app_time = random.randint(2, 5)
        time.sleep(open_app_time)
        driver.background_app(-1)
        driver.terminate_app(self.config.get_app_bundle_id())
        # 再次打开
        # time.sleep(open_app_time)
        # driver.activate_app(self.config.get_app_bundle_id())
        # time.sleep(open_app_time)
        # driver.background_app(-1)
        # driver.terminate_app(self.config.get_app_bundle_id())

        # 关闭代理
        self.close_proxyIp(driver)
        # 还原广告标识符
        self.reduction_idfa(driver)

        # 记录信息
        now_time = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())
        iphone_name = driver.capabilities["deviceName"]
        iphone_udid = driver.capabilities["udid"]
        wb = load_workbook("./serverlogs/tmp.xlsx")
        ws = wb["下载情况表"]
        ws.append([iphone_name, iphone_udid, idfaStr, open_app_time, ip_address, r_status,
                   now_time])
        wb.save("./serverlogs/tmp.xlsx")
        print('---------------------------------------------{}'.format(now_time))
        rows = ws.max_row - 1
        if rows <= 5:
            driver.launch_app()
            self.startTest(driver)

    def proxyIp_Change(self, driver: webdriver, ip_address, ip_port):
        print("[%s]正在设置代理IP" % driver.capabilities["deviceName"])
        driver.activate_app('com.apple.Preferences')
        time.sleep(2)
        driver.find_element(by=By.NAME, value='无线局域网').click()
        time.sleep(2)
        driver.find_element(by=By.NAME, value="更多信息").click()
        time.sleep(1)
        driver.execute_script('mobile: swipe', {'direction': 'up'})
        driver.find_element(by=By.NAME, value='配置代理').click()
        driver.find_element(by=By.NAME, value='手动').click()
        addressFiled = driver.find_element(by=By.NAME, value="服务器")
        addressFiled.clear()
        addressFiled.send_keys(ip_address)
        time.sleep(1)
        porTFiled = driver.find_element(by=By.NAME, value="端口")
        porTFiled.clear()
        porTFiled.send_keys(ip_port)
        driver.find_element(by=By.NAME, value="存储").click()
        print("[%s]设置代理IP成功" % driver.capabilities["deviceName"])
        time.sleep(1)
        driver.background_app(-1)
        driver.terminate_app('com.apple.Preferences')

    def close_proxyIp(self, driver: webdriver):
        print("[%s]正在关闭代理IP" % driver.capabilities["deviceName"])
        driver.activate_app('com.apple.Preferences')
        time.sleep(2)
        driver.find_element(by=By.NAME, value='无线局域网').click()
        time.sleep(2)
        driver.find_element(by=By.NAME, value="更多信息").click()
        time.sleep(1)
        driver.execute_script('mobile: swipe', {'direction': 'up'})
        driver.find_element(by=By.NAME, value='配置代理').click()
        driver.find_element(by=By.NAME, value='关闭').click()
        driver.find_element(by=By.NAME, value="存储").click()
        print("[%s]关闭代理IP成功" % driver.capabilities["deviceName"])
        time.sleep(1)
        driver.back()
        time.sleep(1)
        driver.back()
        time.sleep(1)
        driver.back()

    # 还原广告标识符
    def reduction_idfa(self, driver: webdriver):
        # 隐私
        driver.activate_app('com.apple.Preferences')
        driver.execute_script('mobile: swipe', {'direction': 'up'})
        driver.find_element(by=By.NAME, value='隐私').click()
        driver.execute_script('mobile: swipe', {'direction': 'up'})
        driver.find_element(by=By.NAME, value='广告').click()
        driver.find_element(by=By.NAME, value='还原广告标识符…').click()
        time.sleep(1)
        # 允许弹框
        driver.switch_to.alert.accept()
        time.sleep(1)
        driver.terminate_app('com.apple.Preferences')
        driver.background_app(-1)

    # 打开APP时候的权限处理
    def always_allow(self, driver, number=3):
        for i in range(number):
            loc = ("name", "不允许")
            try:
                e = WebDriverWait(driver, 1, 0.5).until(exc.presence_of_element_located(loc))
                e.click()
            except:
                pass
